from parser_products import get_categories, get_products_info, get_products_hrefs
import openpyxl
import requests


def remove_all_sheets(path: str) -> None:
    wb = openpyxl.load_workbook(filename=path,)
    print(wb.sheetnames)
    for sheet_name in wb.sheetnames:
        wb.remove(wb[sheet_name])
    wb.create_sheet(str(1))
    wb.save(path)


def remove_first_sheet(path: str) -> None:
    wb = openpyxl.load_workbook(filename=path,)
    wb.remove(wb[wb.sheetnames[0]])
    wb.save(path)


def get_sheets_names(path: str):
    wb = openpyxl.load_workbook(filename=path)
    res = wb.sheetnames
    wb.close()
    return res


def write_cat(path: str, cat: dict) -> None:
    wb = openpyxl.load_workbook(filename=path)
    ws = wb.create_sheet(str(cat["category"]).replace('/', ' '))

    ws.cell(row=1, column=1).value = "Категория"
    ws.cell(row=1, column=2).value = "Артикул"
    ws.cell(row=1, column=3).value = "Бренд"
    ws.cell(row=1, column=4).value = "Наименование товара"
    ws.cell(row=1, column=5).value = "Цена"
    ws.cell(row=1, column=6).value = "Описание"
    ws.cell(row=1, column=7).value = "Ссылки на изображения через запятую"

    row = 2
    for product in cat["prods"]:
        ws.cell(row=row, column=1).value = cat["category"]
        ws.cell(row=row, column=2).value = product["art"]
        ws.cell(row=row, column=3).value = product["brand"]
        ws.cell(row=row, column=4).value = product["name_product"]
        ws.cell(row=row, column=5).value = product["price"]
        ws.cell(row=row, column=6).value = product["preview"]
        ws.cell(row=row, column=7).value = product["images"]
        row += 1

    wb.save(path)


def main():
    path_excel = 'product_catalog.xlsx'
    path = "https://yacht-parts.ru"
    products = []
    brands = {}
    with requests.Session() as session:

        # remove_all_sheets(path_excel)

        get_categories(path, products, session)
        sheets = get_sheets_names(path_excel)
        for cat in products:
            category = cat["category"]
            if not ((cat["category"] in sheets) or (str(category).replace('/', ' ')) in sheets):
                get_products_hrefs(path, cat, session)
                for product in cat["prods"]:
                    get_products_info(path, product, brands, session)
                write_cat(path_excel, cat)

        remove_first_sheet(path_excel)
    # remove_all_sheets(path_excel)


main()
